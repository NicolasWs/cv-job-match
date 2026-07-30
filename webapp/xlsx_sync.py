"""Build/merge the Job_Matches spreadsheet and upload it to Google Drive.

Dedupe rule: rows are keyed by job ID (column A). Existing rows keep any
user-added annotation columns; incoming data refreshes the standard columns
only for jobs not already present (so hand-edits are never clobbered), and
fills in a score if the existing row has none.
"""

from __future__ import annotations

import io

import gdrive

HEADERS = [
    "Job ID", "Title", "Company", "Location", "Salary", "Contract",
    "Age (days)", "Recruiter", "Sector", "Score", "Reason", "URL", "Notes",
]


def _job_row(job: dict, score: dict | None) -> list:
    return [
        job.get("id", ""), job.get("title", ""), job.get("company", ""),
        job.get("location", ""), str(job.get("salary", "") or ""),
        job.get("contract_type", ""), job.get("age_days", ""),
        job.get("recruiter", ""), job.get("sector", "") or "",
        (score or {}).get("score", ""), (score or {}).get("reason", ""),
        job.get("url", ""), "",
    ]


def build_and_upload(jobs: list[dict], scores: dict, folder_id: str, filename: str) -> dict:
    from openpyxl import Workbook, load_workbook

    existing_rows: dict[str, list] = {}
    existing_file = gdrive.find_child(folder_id, filename)
    if existing_file:
        wb_old = load_workbook(io.BytesIO(gdrive.download_bytes(existing_file["id"])))
        ws_old = wb_old.active
        rows = list(ws_old.iter_rows(values_only=True))
        if rows and list(rows[0])[: len(HEADERS)] == HEADERS[: len(rows[0])]:
            for row in rows[1:]:
                if row and row[0]:
                    existing_rows[str(row[0])] = list(row)

    wb = Workbook()
    ws = wb.active
    ws.title = "Job Matches"
    ws.append(HEADERS)
    ws.freeze_panes = "A2"

    added = updated = kept = 0
    seen: set[str] = set()
    for job in jobs:
        jid = str(job.get("id", ""))
        seen.add(jid)
        score = scores.get(jid)
        if jid in existing_rows:
            row = existing_rows[jid]
            row += [""] * (len(HEADERS) - len(row))
            if score and row[9] in (None, ""):  # fill score only if user has none
                row[9], row[10] = score.get("score", ""), score.get("reason", "")
                updated += 1
            ws.append(row)
        else:
            ws.append(_job_row(job, score))
            added += 1
    # Keep rows for jobs no longer in this week's table (user may still be tracking them).
    for jid, row in existing_rows.items():
        if jid not in seen:
            ws.append(row + [""] * (len(HEADERS) - len(row)))
            kept += 1

    for col, width in zip(ws.columns, (14, 40, 24, 24, 16, 14, 10, 20, 16, 8, 50, 40, 30)):
        ws.column_dimensions[col[0].column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    file_id = gdrive.upload_bytes(
        folder_id, filename, buf.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return {"file_id": file_id, "added": added, "score_filled": updated, "kept_stale": kept}
